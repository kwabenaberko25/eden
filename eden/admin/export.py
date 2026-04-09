"""
Eden Admin Panel — Export Utilities

Provides CSV, JSON, and Excel export functionality for admin bulk actions.

**Usage:**

    from eden.admin.export import to_csv, to_json, to_excel

    csv_data = await to_csv(records, model)
    json_data = await to_json(records, model)
    excel_bytes = await to_excel(records, model)
"""

import io
import csv
import json
from datetime import datetime, date
from typing import List, Dict, Any, Type, Optional
from decimal import Decimal
from enum import Enum


def _serialize_value(value: Any) -> Any:
    """
    Serialize a Python value to a JSON-compatible type.
    
    Handles:
    - datetime/date objects → ISO format strings
    - Decimal → float
    - Enum → value
    - Objects with __str__ → string representation
    """
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, (list, dict, str, int, float, bool)):
        return value
    # Fallback: convert to string
    return str(value)


def _get_model_fields(model: Type) -> List[str]:
    """
    Extract field names from a SQLAlchemy model.
    
    Returns column names in order of definition.
    """
    try:
        from sqlalchemy import inspect as sa_inspect
        mapper = sa_inspect(model)
        return [col.key for col in mapper.columns]
    except Exception:
        return []


def _get_record_values(record: Any, fields: List[str]) -> Dict[str, Any]:
    """
    Extract field values from a record instance.
    
    Handles missing fields gracefully.
    """
    result = {}
    for field in fields:
        value = getattr(record, field, None)
        result[field] = _serialize_value(value)
    return result


async def to_csv(
    records: List[Any],
    model: Optional[Type] = None,
    fields: Optional[List[str]] = None,
) -> str:
    """
    Convert records to CSV format.
    
    Args:
        records: List of model instances to export.
        model: The model class (used to extract fields if not provided).
        fields: Explicit list of field names to include. If None, auto-detected from model.
    
    Returns:
        CSV string with headers and data rows.
    
    Example:
        csv_str = await to_csv(users, User)
        # Returns: "id,email,name,created_at\n1,user@example.com,John,2024-01-15T10:30:00..."
    """
    if not records:
        return ""
    
    # Determine fields
    if fields is None:
        fields = _get_model_fields(model) if model else list(_get_record_values(records[0], []).keys())
    
    # Build CSV
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()
    
    for record in records:
        row = _get_record_values(record, fields)
        writer.writerow(row)
    
    return output.getvalue()


async def to_json(
    records: List[Any],
    model: Optional[Type] = None,
    fields: Optional[List[str]] = None,
    pretty: bool = True,
) -> str:
    """
    Convert records to JSON format.
    
    Args:
        records: List of model instances to export.
        model: The model class (used to extract fields if not provided).
        fields: Explicit list of field names to include. If None, all fields exported.
        pretty: If True, format with indentation for readability.
    
    Returns:
        JSON string representing the records as an array of objects.
    
    Example:
        json_str = await to_json(users, User)
        # Returns: '[{"id": 1, "email": "user@example.com", ...}, ...]'
    """
    if not records:
        return "[]"
    
    # Determine fields
    if fields is None:
        fields = _get_model_fields(model) if model else None
    
    # Build JSON
    data = []
    for record in records:
        row = _get_record_values(record, fields or list(record.__dict__.keys()))
        data.append(row)
    
    return json.dumps(data, indent=2 if pretty else None, default=str)


async def to_excel(
    records: List[Any],
    model: Optional[Type] = None,
    fields: Optional[List[str]] = None,
    sheet_name: str = "Data",
) -> bytes:
    """
    Convert records to Excel (XLSX) format.
    
    Args:
        records: List of model instances to export.
        model: The model class (used to extract fields if not provided).
        fields: Explicit list of field names to include. If None, auto-detected.
        sheet_name: Name of the worksheet (default: "Data").
    
    Returns:
        Excel file as bytes (XLSX format).
    
    Raises:
        ImportError: If openpyxl is not installed.
    
    Example:
        excel_bytes = await to_excel(users, User)
        # Returns: bytes representing an .xlsx file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )
    
    if not records:
        # Return empty workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    # Determine fields
    if fields is None:
        fields = _get_model_fields(model) if model else list(_get_record_values(records[0], []).keys())
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write header row with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, field_name in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = field_name.replace("_", " ").title()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        row_data = _get_record_values(record, fields)
        for col_idx, field_name in enumerate(fields, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data[field_name]
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for col_idx, field_name in enumerate(fields, start=1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = max(
            len(field_name) + 2,
            20  # Minimum width
        )
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def generate_export_filename(
    model_name: str,
    format: str = "csv",
    timestamp: bool = True,
) -> str:
    """
    Generate a standardized export filename.
    
    Args:
        model_name: Name of the model being exported (e.g., "users", "products").
        format: File extension (csv, json, xlsx).
        timestamp: If True, include ISO timestamp in filename.
    
    Returns:
        Filename string (e.g., "users_2024-01-15T10-30-00.csv").
    
    Example:
        filename = await generate_export_filename("users", "csv")
        # Returns: "users_2024-01-15T10-30-00.csv"
    """
    format = format.lower()
    if format == "excel":
        format = "xlsx"
    
    base = model_name.lower().replace(" ", "_")
    
    if timestamp:
        ts = datetime.now().isoformat().replace(":", "-")[:19]
        return f"{base}_{ts}.{format}"
    
    return f"{base}.{format}"


async def get_export_response_headers(
    filename: str,
    format: str = "csv",
) -> Dict[str, str]:
    """
    Generate HTTP response headers for file downloads.
    
    Args:
        filename: Name of the file to download.
        format: File format (csv, json, xlsx).
    
    Returns:
        Dictionary of headers including Content-Type and Content-Disposition.
    
    Example:
        headers = await get_export_response_headers("users.csv", "csv")
        # Returns: {
        #     "Content-Type": "text/csv; charset=utf-8",
        #     "Content-Disposition": 'attachment; filename="users.csv"'
        # }
    """
    mime_types = {
        "csv": "text/csv; charset=utf-8",
        "json": "application/json; charset=utf-8",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xls": "application/vnd.ms-excel",
    }
    
    format_lower = format.lower()
    content_type = mime_types.get(format_lower, "application/octet-stream")
    
    return {
        "Content-Type": content_type,
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache",
        "Expires": "0",
    }
